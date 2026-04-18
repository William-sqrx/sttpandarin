"""Parse HSK listening exam Excel into a structured form.

Expected layout (Questions sheet):
- Row 1 header: "No" | "Paper 1" | "" | "Paper 2" | "" | ... (each paper takes 2 columns)
- Rows with No 1-5: Part 1 (first column of paper = statement)
- Rows with No 6-10: Part 2 (first column of paper = question)
- Rows with No 11-15: Part 3 (col1 = female Q, col2 = male A)
- Rows with No 16-20: Part 4 (col1 = male statement, col2 = female question)

Exam Instructions sheet: each non-empty cell (in column B) is treated as one instruction
line, read by the female voice.
"""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class QuestionItem:
    paper: str          # e.g. "Paper 1"
    number: int         # 1..20
    part: int           # 1..4
    text_a: str         # first-speaker text
    text_b: str         # second-speaker text (empty for parts 1/2)


@dataclass
class InstructionItem:
    index: int
    text: str


def _part_for_number(n: int) -> int:
    if 1 <= n <= 5:
        return 1
    if 6 <= n <= 10:
        return 2
    if 11 <= n <= 15:
        return 3
    if 16 <= n <= 20:
        return 4
    raise ValueError(f"row number {n} outside 1..20")


def _detect_paper_columns(ws) -> list[tuple[str, int]]:
    """Scan row 1 for 'Paper N' headers. Returns list of (name, 1-indexed col)."""
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if not val:
            continue
        s = str(val).strip()
        if s.lower().startswith("paper"):
            headers.append((s, col))
    return headers


def parse_questions(xlsx_path: Path, sheet_name: str | None = None) -> list[QuestionItem]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Prefer a sheet called "Questions" if present, else the first sheet.
        name = next((n for n in wb.sheetnames if n.strip().lower() == "questions"), None)
        ws = wb[name] if name else wb[wb.sheetnames[0]]

    papers = _detect_paper_columns(ws)
    if not papers:
        raise ValueError("No 'Paper N' headers found in row 1")

    items: list[QuestionItem] = []
    for row in range(2, ws.max_row + 1):
        no_cell = ws.cell(row=row, column=1).value
        if not isinstance(no_cell, (int, float)):
            continue
        n = int(no_cell)
        if n < 1 or n > 20:
            continue
        part = _part_for_number(n)
        for paper_name, col in papers:
            a = ws.cell(row=row, column=col).value
            b = ws.cell(row=row, column=col + 1).value
            a = str(a).strip() if a is not None else ""
            b = str(b).strip() if b is not None else ""
            if part in (1, 2):
                if not a:
                    continue
                items.append(QuestionItem(paper_name, n, part, a, ""))
            else:
                if not a and not b:
                    continue
                items.append(QuestionItem(paper_name, n, part, a, b))
    return items


def parse_instructions(xlsx_path: Path, sheet_name: str | None = None) -> list[InstructionItem]:
    wb = load_workbook(xlsx_path, data_only=True)
    # Try a few common names
    candidates = [sheet_name] if sheet_name else [
        "Exam Instructions", "Exam Instructions Example", "Instructions",
    ]
    ws = None
    for name in candidates:
        if name and name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        return []

    items: list[InstructionItem] = []
    idx = 0
    for row in range(1, ws.max_row + 1):
        # Concatenate all text cells on the row (skip column A if it's a number/label)
        parts: list[str] = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            # Skip pure-number index cells in column A
            if col == 1 and s.isdigit():
                continue
            parts.append(s)
        if not parts:
            continue
        text = " ".join(parts)
        # Filter out pure header labels
        if text.lower() in ("instructions", "exam instructions"):
            continue
        idx += 1
        items.append(InstructionItem(idx, text))
    return items
